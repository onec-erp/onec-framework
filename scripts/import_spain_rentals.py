#!/usr/bin/env python3
"""
One-off importer that loads ~/Downloads/Spain Accounts & Bookings 2022 - 2023.xlsx
into a running OneC example instance via the framework's REST API.

Run the example app first (default http://localhost:8080, admin:admin), then:
    python3 scripts/import_spain_rentals.py [--xlsx PATH] [--limit-bookings N]
"""
import argparse
import os
import sys
from datetime import datetime, date
from typing import Any, Optional

import openpyxl
import requests

DEFAULT_XLSX = os.path.expanduser("~/Downloads/Spain Accounts & Bookings 2022 - 2023.xlsx")
DEFAULT_BASE = "http://localhost:8080"
DEFAULT_AUTH = ("admin", "admin")

# Map spreadsheet labels to enum names (the Java enum constants).
DOC_TYPE_MAP = {
    "documento nacional de identidad": "NATIONAL_ID",
    "documento de identidad": "NATIONAL_ID",
    "dni": "NATIONAL_ID",
    "pasaporte": "PASSPORT",
    "passport": "PASSPORT",
    "permiso de conducir": "DRIVING_LICENSE",
    "driving license": "DRIVING_LICENSE",
}

GENDER_MAP = {"M": "MALE", "F": "FEMALE", "MALE": "MALE", "FEMALE": "FEMALE"}

CHANNEL_MAP = {
    "airbnb": "AIRBNB",
    "booking": "BOOKING_COM",
    "booking.com": "BOOKING_COM",
    "vrbo": "VRBO",
    "homeaway": "VRBO",
    "agency": "AGENCY",
    "direct": "DIRECT",
    "owner": "DIRECT",
}


def s(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return str(v)


def n(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def i(v: Any) -> Optional[int]:
    f = n(v)
    return int(f) if f is not None else None


def to_date_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    try:
        return datetime.fromisoformat(str(v)).date().isoformat()
    except ValueError:
        return None


def to_datetime_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return datetime.combine(v, datetime.min.time()).isoformat()
    return None


class Api:
    def __init__(self, base: str, auth):
        self.base = base.rstrip("/")
        self.session = requests.Session()
        self.session.auth = auth

    def metadata(self) -> dict:
        r = self.session.get(f"{self.base}/api/ui/metadata")
        r.raise_for_status()
        return r.json()

    def list_catalog(self, name: str) -> list:
        r = self.session.get(f"{self.base}/api/ui/catalogs/{name}")
        r.raise_for_status()
        return r.json()

    def post_catalog(self, name: str, body: dict) -> Optional[dict]:
        r = self.session.post(f"{self.base}/api/ui/catalogs/{name}", json=body)
        if not r.ok:
            print(f"  ! catalog {name} failed: {r.status_code} {r.text[:200]}", file=sys.stderr)
            return None
        return r.json()

    def post_document(self, name: str, body: dict) -> Optional[dict]:
        r = self.session.post(f"{self.base}/api/ui/documents/{name}", json=body)
        if not r.ok:
            print(f"  ! document {name} failed: {r.status_code} {r.text[:200]}", file=sys.stderr)
            return None
        return r.json()

    def post_post(self, name: str, doc_id: str) -> bool:
        r = self.session.post(f"{self.base}/api/ui/documents/{name}/{doc_id}/post")
        if not r.ok:
            print(f"  ! post {name}/{doc_id} failed: {r.status_code} {r.text[:200]}", file=sys.stderr)
            return False
        return True


def build_enum_lookup(metadata: dict) -> dict:
    """Walks the metadata manifest collecting enum-value UUIDs by enum name."""
    out: dict[str, dict[str, str]] = {}
    for kind in ("catalogs", "documents"):
        for d in metadata.get(kind, []):
            for a in d.get("attributes", []):
                if a.get("isEnum"):
                    en = a.get("enumName")
                    if en and en not in out:
                        out[en] = {v["name"]: v["id"] for v in a.get("enumValues", [])}
    return out


def load_existing(api: Api, catalog: str, key: str = "_description") -> dict:
    """Returns {description -> _id} for an existing catalog so re-running doesn't duplicate."""
    out = {}
    try:
        for row in api.list_catalog(catalog):
            d = row.get(key)
            if d:
                out[str(d).strip()] = row.get("_id")
    except Exception:
        pass
    return out


# ---------------------------- importers ----------------------------


def import_countries(api: Api, ws) -> dict:
    """Returns {label -> uuid} where label includes English name, Spanish name, and nationality forms."""
    print("→ Countries")
    existing = load_existing(api, "countries")
    by_label: dict[str, str] = {}
    seen_eng: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        eng = s(row[0])
        if not eng or eng in seen_eng:
            continue
        seen_eng.add(eng)
        nat_en = s(row[1])
        spanish_nat = s(row[2])
        spanish_country = s(row[3])

        if eng in existing:
            uuid = existing[eng]
        else:
            res = api.post_catalog("countries", {
                "description": eng,
                "name": eng,
                "nationality": nat_en or "",
            })
            if not res:
                continue
            uuid = res["_id"]

        by_label[eng] = uuid
        if spanish_country:
            by_label[spanish_country] = uuid
        if spanish_nat:
            by_label[spanish_nat] = uuid
        if nat_en:
            by_label[nat_en] = uuid

    print(f"  {len(seen_eng)} unique countries")
    return by_label


def import_bank_accounts(api: Api, ws) -> dict:
    print("→ Bank accounts")
    existing = load_existing(api, "bankaccounts")
    by_iban: dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        iban = s(row[14])
        nominee = s(row[15]) or "Owner"
        if not iban or iban in by_iban:
            continue
        desc = f"{nominee} • {iban[-4:]}"
        if desc in existing:
            by_iban[iban] = existing[desc]
            continue
        res = api.post_catalog("bankaccounts", {
            "description": desc,
            "iban": iban,
            "nominee": nominee,
            "bankName": "",
        })
        if res:
            by_iban[iban] = res["_id"]

    print(f"  {len(by_iban)} accounts")
    return by_iban


def import_properties(api: Api, bookings_ws) -> dict:
    print("→ Properties")
    existing = load_existing(api, "properties")
    by_name: dict[str, str] = {}
    seen: set[str] = set()
    for row in bookings_ws.iter_rows(min_row=2, values_only=True):
        prop = s(row[1])
        if not prop or prop in seen:
            continue
        seen.add(prop)
        if prop in existing:
            by_name[prop] = existing[prop]
            continue
        res = api.post_catalog("properties", {
            "description": prop,
            "displayName": prop,
            "address": "",
        })
        if res:
            by_name[prop] = res["_id"]
    print(f"  {len(by_name)} properties: {sorted(by_name)}")
    return by_name


def map_doc_type(label: Optional[str], enums: dict) -> Optional[str]:
    if not label:
        return None
    key = label.lower()
    name = None
    for fragment, enum_name in DOC_TYPE_MAP.items():
        if fragment in key:
            name = enum_name
            break
    if name is None:
        name = "OTHER"
    return enums.get("Document Types", {}).get(name)


def map_gender(label: Optional[str], enums: dict) -> Optional[str]:
    if not label:
        return None
    name = GENDER_MAP.get(label.upper())
    if not name:
        return None
    return enums.get("Genders", {}).get(name)


def map_channel(label: Optional[str], enums: dict) -> Optional[str]:
    if not label:
        return enums.get("Booking Channels", {}).get("DIRECT")
    key = label.lower()
    for fragment, enum_name in CHANNEL_MAP.items():
        if fragment in key:
            return enums.get("Booking Channels", {}).get(enum_name)
    return enums.get("Booking Channels", {}).get("OTHER")


def import_clients_and_bookings(api: Api, bookings_ws, properties: dict, countries: dict,
                                 enums: dict, year_min: int, year_max: int, limit: int) -> dict:
    print(f"→ Clients + Bookings (years {year_min}–{year_max}, limit {limit})")
    bs_status = enums.get("Booking Statuses", {})
    client_lookup: dict[tuple, str] = {}
    bookings_created = 0
    clients_created = 0

    for row in bookings_ws.iter_rows(min_row=2, values_only=True):
        check_in = row[20]
        if not isinstance(check_in, datetime):
            continue
        if check_in.year < year_min or check_in.year > year_max:
            continue
        if bookings_created >= limit:
            break

        check_out = row[21]
        property_name = s(row[1])
        property_id = properties.get(property_name) if property_name else None
        first_name = s(row[3])
        last1 = s(row[4])
        last2 = s(row[5])
        nat_label = s(row[7])
        doc_no = s(row[8])

        # client dedup: prefer doc number, fall back to name
        client_key = (doc_no.lower() if doc_no else None) or (
            f"{(first_name or '').lower()}|{(last1 or '').lower()}|{(last2 or '').lower()}"
        )

        if client_key in client_lookup:
            primary_client_id = client_lookup[client_key]
        else:
            client_body = {
                "description": " ".join(filter(None, [first_name, last1, last2])) or "(no name)",
                "firstName": first_name or "",
                "lastName1": last1 or "",
                "lastName2": last2 or "",
                "docNumber": doc_no or "",
                "address": s(row[13]) or "",
                "city": s(row[15]) or "",
                "email": s(row[48]) or "",
                "mobile": s(row[49]) or "",
            }
            doc_type_id = map_doc_type(s(row[9]), enums)
            if doc_type_id:
                client_body["docType"] = doc_type_id
            gender_id = map_gender(s(row[11]), enums)
            if gender_id:
                client_body["gender"] = gender_id
            issued = to_date_str(row[10])
            if issued:
                client_body["docIssuedOn"] = issued
            birthday = to_date_str(row[12])
            if birthday:
                client_body["birthday"] = birthday
            if nat_label and nat_label in countries:
                client_body["nationality"] = countries[nat_label]
            country_label = s(row[16])
            if country_label and country_label in countries:
                client_body["country"] = countries[country_label]

            res = api.post_catalog("clients", client_body)
            if not res:
                continue
            primary_client_id = res["_id"]
            client_lookup[client_key] = primary_client_id
            clients_created += 1

        # booking
        status_label = s(row[2])
        canceled = row[22]
        if canceled:
            status_name = "CANCELED"
        elif status_label and status_label.upper() in {"OK", "CONFIRMED"}:
            status_name = "CONFIRMED"
        else:
            status_name = "DRAFT"

        booking_body = {
            "date": to_datetime_str(check_in),
            "property": property_id,
            "status": bs_status.get(status_name),
            "channel": map_channel(s(row[30]), enums),
            "checkIn": to_date_str(check_in),
            "checkOut": to_date_str(check_out),
            "adults": i(row[17]) or 0,
            "children": i(row[18]) or 0,
            "nightRate": n(row[31]) or 0,
            "cleaningFee": n(row[45]) or 0,
            "notes": s(row[52]) or "",
            "primaryClient": primary_client_id,
            "guests": [
                {"client": primary_client_id, "mainGuest": True, "isChild": False}
            ],
        }
        booking_body = {k: v for k, v in booking_body.items() if v is not None and v != ""}
        # Keep zeros; only strip None/empty-string above
        if "adults" not in booking_body:
            booking_body["adults"] = 0
        if "children" not in booking_body:
            booking_body["children"] = 0

        res = api.post_document("bookings", booking_body)
        if not res:
            continue
        bookings_created += 1
        if status_name in ("CONFIRMED",):
            api.post_post("bookings", res["_id"])
        if bookings_created % 25 == 0:
            print(f"  … {bookings_created} bookings, {clients_created} clients")

    print(f"  {clients_created} clients, {bookings_created} bookings")
    return client_lookup


def import_bills(api: Api, bills_ws, properties: dict, year_min: int, year_max: int, limit: int):
    print(f"→ Bills (years {year_min}–{year_max}, limit {limit})")
    created = 0
    # find existing client lookup from server: name -> uuid
    existing_clients = {row.get("_description"): row.get("_id")
                        for row in api.list_catalog("clients") if row.get("_description")}

    for row in bills_ws.iter_rows(min_row=3, values_only=True):
        if created >= limit:
            break
        bill_date = row[2]
        if not isinstance(bill_date, datetime):
            continue
        if bill_date.year < year_min or bill_date.year > year_max:
            continue
        client_name = s(row[5])
        gross = n(row[11])
        net = n(row[12])
        iva_pct = n(row[14])
        if gross is None or net is None:
            continue

        client_id = existing_clients.get(client_name) if client_name else None
        if client_id is None:
            # Create a thin client record on-the-fly from bill info
            res = api.post_catalog("clients", {
                "description": client_name or "(unknown)",
                "firstName": (client_name or "").split(" ", 1)[0] if client_name else "",
                "lastName1": (client_name or "").split(" ", 1)[1] if client_name and " " in client_name else "",
                "docNumber": s(row[6]) or "",
                "address": s(row[7]) or "",
                "city": s(row[9]) or "",
                "postCode": s(row[8]) or "",
                "email": "",
                "mobile": "",
            })
            if not res:
                continue
            client_id = res["_id"]
            if client_name:
                existing_clients[client_name] = client_id

        body = {
            "date": to_datetime_str(bill_date),
            "client": client_id,
            "net": round(float(net), 2),
            "ivaPercent": round(float(iva_pct or 10), 2),
            "comments": "",
        }
        res = api.post_document("bills", body)
        if not res:
            continue
        created += 1
        api.post_post("bills", res["_id"])
        if created % 25 == 0:
            print(f"  … {created} bills")

    print(f"  {created} bills")


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", default=DEFAULT_XLSX)
    p.add_argument("--base", default=DEFAULT_BASE)
    p.add_argument("--user", default=DEFAULT_AUTH[0])
    p.add_argument("--password", default=DEFAULT_AUTH[1])
    p.add_argument("--year-min", type=int, default=2022)
    p.add_argument("--year-max", type=int, default=2023)
    p.add_argument("--limit-bookings", type=int, default=200)
    p.add_argument("--limit-bills", type=int, default=150)
    args = p.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"ERROR: spreadsheet not found at {args.xlsx}", file=sys.stderr)
        return 1

    api = Api(args.base, (args.user, args.password))
    print(f"Using {args.base} as {args.user}")
    md = api.metadata()
    enums = build_enum_lookup(md)
    print(f"Enums loaded: {sorted(enums)}")

    print(f"Loading {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx, data_only=True, read_only=False)

    countries = import_countries(api, wb["Data"])
    banks = import_bank_accounts(api, wb["Data"])
    properties = import_properties(api, wb["Bookings"])
    import_clients_and_bookings(api, wb["Bookings"], properties, countries, enums,
                                args.year_min, args.year_max, args.limit_bookings)
    import_bills(api, wb["Bills"], properties, args.year_min, args.year_max, args.limit_bills)

    print("✓ Import complete")
    return 0


if __name__ == "__main__":
    sys.exit(main())
